#!/usr/bin/env python3
"""
Apply sds-research.js (and sds-verify.js) results back into the review xlsx.

Usage:
    python3 apply_results.py <xlsx_path> <sheet_name> <results_json> <row_mapping_json> \
        [--column-map column_map.json] [--memo-col memo]

results_json: the `results` (or `final_results`, from sds-verify.js) array. Each item has:
    row, product_name, resolution ("found"|"formulation"|"not_formulation"|"unresolved"),
    field_values ({...depending on which fields were requested...}),
    not_formulation_reason, sds_summary ({...completeness checklist + summary_ko...}),
    source_url, confidence, notes

row_mapping_json: output of extract_targets.py (product_name -> [row, ...])

Column mapping
--------------
formulation_code has a built-in default mapping to this sheet's existing
formulation_type / formulation_code / formulation_type_ko / not_formulation_reason columns
(the columns this harness was originally built around).

Every OTHER field (ingredients, cas_number, physicochemical, toxicity) has NO destination
column by default, since those columns don't exist in the sheet yet - per the current scope
decision, this script only adds the *capability* to collect and record that data; it is not
responsible for designing the eventual spreadsheet schema. Pass --column-map to point a field
at a real column once one exists, e.g.:
    {"toxicity": {"col": "toxicity_info"}, "cas_number": {"col": "cas_number"}}
A field with no mapping (or whose mapped column doesn't exist in the sheet) is NOT dropped -
its value and the sds_summary completeness line are recorded as an Excel comment on the
product_name cell instead, so nothing collected is silently lost while the schema is still
undecided.

Regardless of which field(s) were requested, source_url (if present) always overwrites
formulation_src and is inserted - in blue font - into the first EMPTY column among
[tox_source_url, ingredient_source, supplemental_source_url]. Existing values are never
overwritten, even if a mismatch was flagged in `notes`.
"""
import argparse
import copy
import json
import os
import sys

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font

BLUE = Font(color="FF0000FF")

DEFAULT_COLUMN_MAP = {
    "formulation_code": {
        "type_col": "formulation_type",
        "code_col": "formulation_code",
        "ko_col": "formulation_type_ko",
        "reason_col": "not_formulation_reason",
    },
}


def load_code_map(harness_dir):
    with open(f"{harness_dir}/cipac_codes.json") as f:
        data = json.load(f)
    m = {}
    for entry in data["current"] + data["discontinued"]:
        m[entry["code"]] = entry["term_ko"]
    return m, {e["code"] for e in data["current"]}, {e["code"] for e in data["discontinued"]}


def apply_formulation_code(ws, idx, row, field_values, resolution, not_formulation_reason, code_to_ko, current_codes, discontinued_codes, unrecognized_codes, colmap):
    type_col, code_col, ko_col, reason_col = colmap["type_col"], colmap["code_col"], colmap["ko_col"], colmap["reason_col"]
    missing = [c for c in (type_col, code_col, ko_col, reason_col) if c not in idx]
    if missing:
        return False  # caller falls back to a comment

    if resolution in ("found", "formulation"):
        fv = (field_values or {}).get("formulation_code") or {}
        code = fv.get("code") if isinstance(fv, dict) else fv
        term_en = fv.get("term_en") if isinstance(fv, dict) else None
        if code:
            base = code.split("-")[0]
            if code not in current_codes and code not in discontinued_codes and base not in current_codes and base not in discontinued_codes:
                unrecognized_codes.add(code)
            ftype_ko = code_to_ko.get(code) or code_to_ko.get(base)
        else:
            ftype_ko = None
        ftype = ftype_ko or term_en
        ws.cell(row=row, column=idx[type_col]).value = ftype
        ws.cell(row=row, column=idx[code_col]).value = code
        ws.cell(row=row, column=idx[ko_col]).value = ftype_ko
        ws.cell(row=row, column=idx[reason_col]).value = None
    elif resolution == "not_formulation":
        ws.cell(row=row, column=idx[type_col]).value = None
        ws.cell(row=row, column=idx[code_col]).value = None
        ws.cell(row=row, column=idx[ko_col]).value = None
        ws.cell(row=row, column=idx[reason_col]).value = not_formulation_reason
    return True


def apply_generic_field(ws, idx, row, field, value, colmap):
    col = (colmap or {}).get("col")
    if not col or col not in idx:
        return False
    cell = ws.cell(row=row, column=idx[col])
    cell.value = json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value
    return True


def append_comment(ws, idx, row, text):
    cell = ws.cell(row=row, column=idx["product_name"])
    existing = cell.comment.text if cell.comment else ""
    combined = (existing + "\n" if existing else "") + text
    cell.comment = Comment(combined[:32000], "sds-research harness")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("sheet_name")
    ap.add_argument("results_json")
    ap.add_argument("row_mapping_json")
    ap.add_argument("--harness-dir", default=None)
    ap.add_argument("--column-map", default=None, help="JSON file: {field: {\"col\": \"...\"}} for fields beyond formulation_code")
    args = ap.parse_args()

    harness_dir = args.harness_dir or os.path.dirname(os.path.abspath(__file__))
    code_to_ko, current_codes, discontinued_codes = load_code_map(harness_dir)

    column_map = dict(DEFAULT_COLUMN_MAP)
    if args.column_map:
        column_map.update(json.load(open(args.column_map)))

    results = json.load(open(args.results_json))
    name_to_rows = json.load(open(args.row_mapping_json))

    expanded = {}
    for r in results:
        rows_for_name = name_to_rows.get(r.get("product_name"), [r.get("row")])
        for row in rows_for_name:
            expanded[row] = r

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=False)
    ws = wb[args.sheet_name]
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    unrecognized_codes = set()
    changed = 0
    skipped_unresolved = 0

    for row, r in sorted(expanded.items()):
        resolution = r.get("resolution")
        source_url = r.get("source_url")
        field_values = r.get("field_values") or {}
        sds_summary = r.get("sds_summary") or {}

        if resolution == "unresolved" or resolution is None:
            skipped_unresolved += 1
            continue

        any_written = False
        if "formulation_code" in column_map:
            wrote = apply_formulation_code(
                ws, idx, row, field_values, resolution, r.get("not_formulation_reason"),
                code_to_ko, current_codes, discontinued_codes, unrecognized_codes,
                column_map["formulation_code"],
            )
            any_written = any_written or wrote
            if not wrote and (field_values.get("formulation_code") or resolution == "not_formulation"):
                append_comment(ws, idx, row, f"[formulation_code] {json.dumps(field_values.get('formulation_code'), ensure_ascii=False)} / reason={r.get('not_formulation_reason')}")

        for field, value in field_values.items():
            if field == "formulation_code":
                continue
            wrote = apply_generic_field(ws, idx, row, field, value, column_map.get(field))
            any_written = any_written or wrote
            if not wrote:
                append_comment(ws, idx, row, f"[{field}] {json.dumps(value, ensure_ascii=False)}")

        if sds_summary:
            summary_ko = sds_summary.get("summary_ko")
            memo_col = column_map.get("_memo", {}).get("col")
            if memo_col and memo_col in idx:
                ws.cell(row=row, column=idx[memo_col]).value = summary_ko
            elif summary_ko:
                append_comment(ws, idx, row, f"[SDS 요약] {summary_ko} (source: {source_url})")

        if any_written or field_values or resolution == "not_formulation":
            changed += 1

        if source_url:
            ws.cell(row=row, column=idx["formulation_src"]).value = source_url
            for col in ["tox_source_url", "ingredient_source", "supplemental_source_url"]:
                cell = ws.cell(row=row, column=idx[col])
                if cell.value is None:
                    cell.value = source_url
                    cell.font = copy.copy(BLUE)
                    break

    wb.save(args.xlsx_path)

    print(f"Rows updated: {changed}")
    print(f"Rows left unresolved (untouched): {skipped_unresolved}")
    if unrecognized_codes:
        print(f"WARNING: codes not found in cipac_codes.json: {sorted(unrecognized_codes)}", file=sys.stderr)


if __name__ == "__main__":
    main()
