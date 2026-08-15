#!/usr/bin/env python3
"""
Split input_dataset.xlsx into per-role team assignment sheets for manual MSDS/SDS
data collection (ingredients / formulation code / toxicity code / physicochemical
properties), in the same style as 2차검토완료_종합.xlsx (bold headers, dark-red
highlighted team-input columns, dropdown validation on the verification column).

Usage:
    python3 build_ml_collection_sheets.py <input_dataset.xlsx> [--out <output.xlsx>]

Reads:
    <input_dataset.xlsx>#formulation  - 1 row per product (전체 컬럼)
    <input_dataset.xlsx>#ingredient   - 1 row per product x ingredient

Writes one workbook with 5 sheets:
    설명        - what this file is, per-sheet columns, 검증여부 값 설명
    성분        - ingredient 시트 기반, 1 row = 1 성분 (전체 행)
    제형코드    - formulation 시트 기반, 1 row = 1 제형 (전체 행)
    독성코드    - formulation 시트 기반, 1 row = 1 제형 (전체 행)
    특성_물리화학적 - formulation 시트 기반, 1 row = 1 제형 (전체 행, 기존 데이터 없음)

Every sheet gets the same 4 team-input columns appended (dark-red header, matching
2차검토완료_종합.xlsx's 팀원 기록 컬럼 style):
    신규소스링크   - 자동 수집 소스가 부실/부적합할 때 팀원이 새로 찾은 더 나은 SDS/라벨 URL
    추출정보       - 그 시트의 역할(성분/제형코드/독성코드/물리화학적특성)에 해당하는,
                     MSDS에서 직접 추출/확인한 값
    메모           - 기타 특이사항
    검증여부       - 드롭다운: 미확인 / 검증완료 / 정보없음 / 재검토필요
"""
import argparse

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

TEAM_FILL = PatternFill("solid", fgColor="FF7A1F1F")
TEAM_FONT = Font(bold=True, color="FFFFFFFF")
BASE_FONT = Font(bold=True)

VERIFY_VALUES = ["미확인", "검증완료", "정보없음", "재검토필요"]

NEW_COLS = ["신규소스링크", "추출정보", "메모", "검증여부"]

SHEET_SPECS = [
    {
        "name": "성분",
        "source_sheet": "ingredient",
        "columns": [
            "Formulation_ID", "product_name", "ing_idx", "ingredient_name", "cas",
            "pct_text", "ingredient_source", "doc_rel", "align_ok",
            "structure_completeness",
        ],
        "extract_label": "추출정보(성분명 | CAS | 함량% - SDS Section 3 기준)",
    },
    {
        "name": "제형코드",
        "source_sheet": "formulation",
        "columns": [
            "Formulation_ID", "product_name", "n_ingredients", "ingredient_names",
            "formulation_type", "formulation_code", "formulation_type_ko",
            "formulation_src", "not_formulation_reason", "ingredient_source",
            "tox_source_url", "doc_source", "doc_rel",
        ],
        "extract_label": "추출정보(CIPAC/CropLife 제형코드, 예: SC/WG/EC)",
    },
    {
        "name": "독성코드",
        "source_sheet": "formulation",
        "columns": [
            "Formulation_ID", "product_name", "ingredient_names", "tox_signal_word",
            "tox_h_statements", "cat_eye_ghs", "cat_skin_ghs", "cat_sens_ghs",
            "sds_grade_status_eye", "sds_grade_status_skin", "sds_grade_status_sens",
            "tox_source_url", "doc_source", "doc_rel",
        ],
        "extract_label": "추출정보(GHS 신호어/H코드/독성 구분 - SDS Section 11 기준)",
    },
    {
        "name": "특성_물리화학적",
        "source_sheet": "formulation",
        "columns": [
            "Formulation_ID", "product_name", "ingredient_names", "ingredient_source",
            "tox_source_url", "doc_source", "doc_rel",
        ],
        "extract_label": "추출정보(외관/pH/비중/인화점/증기압 등 - SDS Section 9 기준, ' | '로 구분)",
    },
]


def read_sheet_rows(src_wb, sheet_name, columns):
    ws = src_wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    missing = [c for c in columns if c not in idx]
    if missing:
        raise SystemExit(f"'{sheet_name}' sheet is missing expected columns: {missing}")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append(tuple(r[idx[c]] for c in columns))
    return rows


def write_sheet(wb, spec, rows):
    ws = wb.create_sheet(spec["name"])
    headers = list(spec["columns"]) + NEW_COLS
    ws.append(headers)

    n_base = len(spec["columns"])
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i)
        if i > n_base:
            cell.font = TEAM_FONT
            cell.fill = TEAM_FILL
        else:
            cell.font = BASE_FONT

    for row in rows:
        ws.append(list(row) + [None, None, None, None])

    last_row = ws.max_row
    verify_col = n_base + len(NEW_COLS)
    verify_letter = get_column_letter(verify_col)
    dv = DataValidation(type="list", formula1='"' + ",".join(VERIFY_VALUES) + '"', allow_blank=True)
    dv.add(f"{verify_letter}2:{verify_letter}{last_row}")
    ws.add_data_validation(dv)

    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        header_len = len(str(col_cells[0].value or ""))
        ws.column_dimensions[letter].width = max(12, min(header_len + 4, 40))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    return ws


def write_readme(wb, counts):
    ws = wb.create_sheet("설명", 0)
    rows = [
        ("MSDS 학습정보 수집 배정 — 설명", None),
        ("", None),
        ("1. 이 파일은 무엇인가", None),
        ("input_dataset.xlsx(제형/성분 전체 데이터)를 역할별로 4개 시트로 나눈 작업 배정 파일입니다.",
         "각 시트를 팀원 1명씩 담당해 MSDS(SDS)에서 해당 정보를 직접 확인/수집합니다."),
        ("", None),
        ("2. 시트 구성", None),
        ("시트", "내용"),
        ("성분", f"성분 1개 = 1행 (총 {counts['성분']}행). ingredient_name/cas/pct_text는 자동 추출된 참고값 — "
                "SDS Section 3과 대조해 맞는지 확인하고, 다르면 추출정보 칸에 정정값을 입력."),
        ("제형코드", f"제형 1개 = 1행 (총 {counts['제형코드']}행). formulation_code가 이미 채워진 행도 SDS/라벨과 "
                  "대조 확인 권장(자동화 오류 가능). 비어있는 행이 최우선 작업 대상."),
        ("독성코드", f"제형 1개 = 1행 (총 {counts['독성코드']}행). tox_signal_word/H코드/GHS 구분이 참고값. "
                  "SDS Section 11 기준으로 신호어·H코드·독성 구분을 추출정보에 기입."),
        ("특성_물리화학적", f"제형 1개 = 1행 (총 {counts['특성_물리화학적']}행). 기존 데이터가 전혀 없는 "
                        "영역 — SDS Section 9(외관/pH/비중/인화점/증기압/수용해도 등)를 처음부터 수집."),
        ("", None),
        ("3. 공통 팀원 입력 컬럼 (진한 빨간 헤더)", None),
        ("컬럼", "내용"),
        ("신규소스링크", "기존 ingredient_source/tox_source_url/doc_rel로 연 SDS가 부실하거나(단일물질뿐, "
                     "섹션 누락 등) 아예 안 열리면, 제조사 사이트/EPA 라벨/다른 SDS 애그리게이터에서 "
                     "더 나은 소스를 찾아 그 URL을 여기에 붙여넣기."),
        ("추출정보", "그 시트 역할에 맞는, MSDS에서 직접 확인한 값. 아래 '4. 시트별 추출정보 형식' 참고."),
        ("메모", "특이사항(예: 성분명이 다름, 로그인 필요, SDS에 해당 섹션 없음, 대체 검색어 등)."),
        ("검증여부", f"드롭다운: {' / '.join(VERIFY_VALUES)}. 확인을 시작하기 전엔 비워두고, 확인 후 선택."),
        ("", None),
        ("4. 시트별 추출정보 형식", None),
        *[(spec["name"], spec["extract_label"]) for spec in SHEET_SPECS],
        ("", None),
        ("5. 검증여부 값 의미", None),
        ("값", "의미"),
        ("미확인", "아직 확인 전 (기본값, 빈 칸과 동일하게 취급)"),
        ("검증완료", "SDS를 직접 확인했고 값이 맞음(또는 정정 완료)"),
        ("정보없음", "SDS를 확인했으나 해당 정보 자체가 문서에 없음"),
        ("재검토필요", "SDS와 값이 어긋나거나 소스를 못 찾음 — 메모에 사유 기록, 최우선 재검토 대상"),
        ("", None),
        ("6. 참고", None),
        ("· ingredient_source/tox_source_url/doc_rel은 input_dataset.xlsx에서 그대로 가져온 참고값입니다 — "
         "지우지 말고 옆의 신규소스링크에 추가하세요.", None),
        ("· doc_rel은 로컬 SDS 첨부파일 상대경로입니다(입력 데이터셋 기준 폴더 참고).", None),
        ("· 이 파일은 formulation_harness/build_ml_collection_sheets.py로 생성되었습니다 — "
         "input_dataset.xlsx가 갱신되면 스크립트를 다시 실행해 재생성하세요.", None),
    ]
    for r in rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 90
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r in (1, 3, 6, 12, 17, 22, 28):
        if r <= ws.max_row:
            ws.cell(row=r, column=1).font = Font(bold=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input_dataset_xlsx")
    ap.add_argument("--out", default="ML학습정보수집_배정.xlsx")
    args = ap.parse_args()

    src_wb = openpyxl.load_workbook(args.input_dataset_xlsx, read_only=True, data_only=True)

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    counts = {}
    for spec in SHEET_SPECS:
        rows = read_sheet_rows(src_wb, spec["source_sheet"], spec["columns"])
        write_sheet(out_wb, spec, rows)
        counts[spec["name"]] = len(rows)

    write_readme(out_wb, counts)
    out_wb.move_sheet("설명", offset=-len(out_wb.sheetnames))

    out_wb.save(args.out)

    print("Sheets written:")
    for name, n in counts.items():
        print(f"  {name}: {n} rows")
    print(f"Wrote: {args.out}")


if __name__ == "__main__":
    main()
