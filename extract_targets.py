#!/usr/bin/env python3
"""
Extract rows needing SDS research from a review sheet.

Usage:
    python3 extract_targets.py <xlsx_path> <sheet_name> \
        [--fields formulation_code,toxicity] [--recheck-reasons other] \
        [--all-rows] [--out-dir .]

Selection logic:
  - Default (no --all-rows): selects rows where formulation_type is empty AND
    (not_formulation_reason is empty OR in --recheck-reasons). This is the only
    field with a real destination column today (see apply_results.py's
    DEFAULT_COLUMN_MAP), so it is the only field this script can currently use
    to decide whether a row is "already done".
  - --all-rows: bypasses the formulation-based filter entirely and selects every
    row. Use this when --fields is purely ingredients/cas_number/physicochemical/
    toxicity and you want the harness to research those for every product
    regardless of formulation status (no destination columns exist yet to check
    "already filled" for those fields).
  - --fields is recorded into unique_products.json as metadata (which role/fields
    you intend to collect) but does not yet change the row-selection logic beyond
    the two options above.

Dedupes by product_name and writes two files into --out-dir:
    unique_products.json  - one entry per unique product_name + the --fields list
                             (for the sds-research.js Workflow's `args`)
    row_mapping.json      - product_name -> [row numbers], to expand results back later
"""
import argparse
import json
import sys
import openpyxl

REQUIRED_COLUMNS = [
    "Formulation_ID", "product_name", "ingredient_names", "doc_source", "doc_rel",
    "tox_doc_product", "tox_source_url", "supplemental_source_url", "ingredient_source",
    "formulation_type", "formulation_code", "formulation_type_ko", "formulation_src",
    "not_formulation_reason",
]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("sheet_name")
    ap.add_argument("--recheck-reasons", default="", help="Comma-separated not_formulation_reason values to re-examine even though already filled (e.g. 'other')")
    ap.add_argument("--fields", default="formulation_code", help="Comma-separated role fields to collect: formulation_code,ingredients,cas_number,physicochemical,toxicity. Recorded as metadata; see module docstring for how each affects row selection today.")
    ap.add_argument("--all-rows", action="store_true", help="Bypass the formulation-based filter and select every row (use for fields with no destination column yet).")
    ap.add_argument("--out-dir", default=".")
    args = ap.parse_args()

    recheck = {r.strip() for r in args.recheck_reasons.split(",") if r.strip()}
    fields = [f.strip() for f in args.fields.split(",") if f.strip()]

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=False)
    ws = wb[args.sheet_name]
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    missing_cols = [c for c in REQUIRED_COLUMNS if c not in idx]
    if missing_cols:
        print(f"WARNING: sheet is missing expected columns: {missing_cols}", file=sys.stderr)

    targets = []
    for r in range(2, ws.max_row + 1):
        if not args.all_rows:
            ftype = ws.cell(row=r, column=idx["formulation_type"]).value if "formulation_type" in idx else None
            reason = ws.cell(row=r, column=idx["not_formulation_reason"]).value if "not_formulation_reason" in idx else None
            if ftype:
                continue
            if reason and reason not in recheck:
                continue
        else:
            reason = ws.cell(row=r, column=idx["not_formulation_reason"]).value if "not_formulation_reason" in idx else None
        rec = {"row": r}
        for col in ["product_name", "ingredient_names", "doc_source", "doc_rel",
                    "tox_doc_product", "tox_source_url", "supplemental_source_url",
                    "ingredient_source", "formulation_src"]:
            if col in idx:
                rec[col] = ws.cell(row=r, column=idx[col]).value
        rec["current_not_formulation_reason"] = reason
        targets.append(rec)

    seen = {}
    unique = []
    for t in targets:
        key = t.get("product_name")
        if key in seen:
            seen[key].append(t["row"])
            continue
        seen[key] = [t["row"]]
        unique.append(t)

    import os
    os.makedirs(args.out_dir, exist_ok=True)
    with open(os.path.join(args.out_dir, "unique_products.json"), "w") as f:
        json.dump(unique, f, ensure_ascii=False, indent=2)
    with open(os.path.join(args.out_dir, "row_mapping.json"), "w") as f:
        json.dump(seen, f, ensure_ascii=False, indent=2)
    with open(os.path.join(args.out_dir, "fields.json"), "w") as f:
        json.dump({"fields": fields}, f, ensure_ascii=False, indent=2)

    print(f"Fields requested: {fields}")
    print(f"Total target rows: {len(targets)}")
    print(f"Unique products: {len(unique)}")
    print(f"Wrote: {args.out_dir}/unique_products.json, {args.out_dir}/row_mapping.json, {args.out_dir}/fields.json")


if __name__ == "__main__":
    main()
